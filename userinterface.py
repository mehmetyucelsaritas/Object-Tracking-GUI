import sys
import matplotlib.pyplot as plt
import openpyxl
import numpy as np
from tkinter import *
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg, NavigationToolbar2Tk


class UserInterface:
    def __init__(self, images, object_coordinates, video_fps, xlsx_path):
        self.window = Tk()
        self.image_fig = None  # matplotlib figure
        self.draw_fig = None  # matplotlib figure
        self.img_subplot = None
        self.draw_canvas = None
        self.img_canvas = None
        self.next_button = None
        self.back_button = None
        self.save_button = None
        self.file_select_button = None
        self.select_frame_entry = None
        self.text_label = None
        self.cage_coordinates = object_coordinates["Cage"]
        self.cage_coordinates_mean = np.mean(self.cage_coordinates)
        self.fish_coordinates = object_coordinates["Fish"]
        self.fish_coordinates_mean = np.mean(self.fish_coordinates)
        self.labelled_fish_coordinates = None
        self.labelled_cage_coordinates = None
        self.current_frame = 0
        self.click_number = 0  # when two click pressed we can not click anymore
        self.video_images = images  # folds all images as list
        self.fps = video_fps
        self.tk_variable = IntVar(self.window, self.current_frame)  # special tkinter int type variable
        self.frame = Frame(self.window)  # Tkinter frame object to locate toolbar
        self.toolbar_show = True  # to show toolbar in window
        self.filename = ""
        self.xlsx_path = xlsx_path
        self.workbook = openpyxl.load_workbook(filename=self.xlsx_path)
        self.worksheet = self.workbook.active  # activate worksheet
        self.time_domain_index = None

    def save_changes2excel(self):
        """
        Overwrites Updated Coordinates to xlsx file.
        Closes Gui and finishes all operations.

        :return: empty
        """
        self.workbook.save(filename=self.xlsx_path)
        self.window.destroy()

    def click(self, event):
        """
        When clicked twice, plots dots on image and gets new Cage and Fish coordinates.

        :param event:
        :return: empty
        """
        try:
            x, y = int(event.xdata), int(event.ydata)

            if event.dblclick and self.click_number < 2:
                if self.click_number == 0:
                    self.img_subplot.plot([x], [y], linestyle="none", marker="o", color="b")
                    self.labelled_fish_coordinates = x
                    self.worksheet.cell(column=1, row=self.current_frame + 2, value=self.labelled_fish_coordinates)
                else:
                    self.img_subplot.plot([x], [y], linestyle="none", marker="o", color="r")
                    self.labelled_cage_coordinates = x
                    self.worksheet.cell(column=2, row=self.current_frame + 2, value=self.labelled_cage_coordinates)
                # print('{}, {}'.format(x, y))
                self.click_number += 1

                self.update_coordinate()
        except:
            print("Please Click on to the image")

    def update_coordinate(self):
        """
        Each time clicked, it updates fish and cage coordinate arrays (i.e self.fish_coordinates, self.cage_coordinates)

        :return: empty
        """
        try:
            if self.labelled_fish_coordinates is not None:
                self.fish_coordinates[self.current_frame] = self.labelled_fish_coordinates
            if self.labelled_cage_coordinates is not None:
                self.cage_coordinates[self.current_frame] = self.labelled_cage_coordinates
        except:
            print("Error occured at line 77")

    def next_frame(self, event=None):
        """
        When the right arrow key is pressed, the next frame is showed

        :param event:
        :return: empty
        """
        try:
            if 0 <= self.current_frame < len(self.video_images) - 1:
                self.current_frame += 1
                self.tk_variable.set(self.current_frame)
        except:
            self.current_frame = self.current_frame  # Do nothing

    def previous_frame(self, event=None):
        """
        When the left arrow key is pressed, the previous frame is showed

        :param event:
        :return: empty
        """
        try:
            if 1 <= self.current_frame < len(self.video_images):
                self.current_frame -= 1
                self.tk_variable.set(self.current_frame)
        except:
            self.current_frame = self.current_frame  # Do nothing

    def set_click_number(self, *args):
        """
        Every time when we change current frame we set click number to zero

        :param args:
        :return:
        """

        self.click_number = 0

    def go_entered_frame(self, *args):
        """
        When entering any number between 0 and max frame number, goes to the entered number.

        :param args:
        :return: empty
        """
        try:
            if 0 <= int(self.select_frame_entry.get()) < len(self.video_images):
                self.current_frame = int(self.select_frame_entry.get())
                self.tk_variable.set(self.current_frame)
        except:
            pass

    def initialize_elements(self):
        """
        Configure and create gui elements.

        :return: empty
        """
        # ------------------ Window --------------------
        self.window.title("NeuroLab Tracking Gui")
        self.window.config(background="white")
        self.window.bind('<Right>', self.next_frame)
        self.window.bind('<Left>', self.previous_frame)
        self.window.bind('<Return>', self.go_entered_frame)

        # ------------------- Figures -------------------
        self.image_fig = plt.figure(figsize=(8.5, 3.7))
        self.image_fig.suptitle("Tracked Video")
        self.img_subplot = self.image_fig.add_subplot(111)
        # self.draw_fig, (self.ax1, self.ax2) = plt.subplots(2, 1, figsize=(5.5, 4))
        self.draw_fig, (self.ax1) = plt.subplots(figsize=(5.5, 3.7))
        self.draw_fig.suptitle("Tracking Analysis")
        self.ax1.set(xlabel="Time", ylabel="Cage and Fish Coordinate")
        # self.ax2.set(xlabel="Time", ylabel="Fish Coordinate")
        # plt.subplots_adjust(wspace=0.4, hspace=0.6)

        # ------------------- Canvases ------------------
        self.draw_canvas = FigureCanvasTkAgg(self.draw_fig, master=self.window)
        self.draw_canvas.draw()
        self.draw_canvas.get_tk_widget().grid(row=0, column=1)
        self.img_canvas = FigureCanvasTkAgg(self.image_fig, master=self.window)
        self.img_canvas.get_tk_widget().grid(row=0, column=0)
        self.img_canvas.mpl_connect('button_press_event', self.click)

        # ------------------- Buttons -------------------
        self.next_button = Button(text="Next", highlightthickness=0, command=self.next_frame)
        self.next_button.grid(column=1, row=2, sticky=E, padx=50)
        self.back_button = Button(text="Back", highlightthickness=0, command=self.previous_frame)
        self.back_button.grid(column=1, row=2, sticky=E, padx=85)
        self.save_button = Button(text="Save Changes", highlightthickness=0, command=self.save_changes2excel)
        self.save_button.grid(column=1, row=3)

        # ------------------- Entry ---------------------
        self.select_frame_entry = Entry(width=6)
        self.select_frame_entry.grid(column=1, row=2, sticky=W, padx=130)

        # ------------------- Label ---------------------
        self.text_label = Label(text="Enter Frame Number: ")
        self.text_label.grid(column=1, row=2, sticky=W)

        # ------------------ ToolBar -------------------
        self.frame.grid(column=0, row=1)
        if self.toolbar_show:
            NavigationToolbar2Tk(self.img_canvas, window=self.frame)

        # ------------------ To Detect If Frame Changed ------------
        self.tk_variable.trace("w", self.set_click_number)

    def show_window(self):
        """
        Draw each figure on Tkinter window and refresh page till tk variable changes.

        :return:empty
        """
        # -------------------- Drawing Cage and Fish Plots on Tkinter --------------------
        self.time_domain_index = self.cage_coordinates.index / self.fps
        self.ax1.plot(self.time_domain_index, self.cage_coordinates.values - self.cage_coordinates_mean, color="r", label="Cage")
        self.ax1.plot(self.time_domain_index, self.fish_coordinates.values - self.fish_coordinates_mean, color="b", label="Fish")
        self.ax1.legend()
        # self.ax2.plot(self.fish_coordinates.index / self.fps, self.fish_coordinates.values, color="r")

        # -------------------- Drawing Fish İmages on Tkinter ----------------------
        while True:
            self.img_subplot.clear()
            self.img_subplot.imshow(self.video_images[self.current_frame])
            self.img_canvas.draw()
            self.back_button.wait_variable(self.tk_variable)
            print(f"Current Frame {self.tk_variable.get()}")
